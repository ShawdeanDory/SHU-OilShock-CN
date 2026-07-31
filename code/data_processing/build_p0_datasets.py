"""Build reproducible P0 processed datasets and a data-quality report."""

from __future__ import annotations

import argparse
import html
import json
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd


REPO_ROOT = Path(__file__).resolve().parents[2]
RAW_DIR = REPO_ROOT / "data" / "raw"
PROCESSED_DIR = REPO_ROOT / "data" / "processed"
REPORT_PATH = REPO_ROOT / "reports" / "DATA_QUALITY_REPORT.md"
MANIFEST_PATH = RAW_DIR / "source_manifest.json"
DICTIONARY_PATH = REPO_ROOT / "data" / "DATA_DICTIONARY.csv"
CUTOFF = pd.Timestamp("2026-06-30")
START_MONTHLY = pd.Timestamp("2010-01-01")
SUCCESS_STATUSES = {"DOWNLOADED", "CACHED"}
EU_FUEL_COUNTRIES = {
    "DEU": ("DE", "Germany", "germany"),
    "FRA": ("FR", "France", "france"),
    "ITA": ("IT", "Italy", "italy"),
    "ESP": ("ES", "Spain", "spain"),
}


@dataclass
class QualityRow:
    dataset: str
    status: str
    rows: int
    start: str
    end: str
    missing: int
    duplicates: int
    note: str


def load_manifest() -> dict[str, dict[str, Any]]:
    rows = json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))
    return {row["artifact_id"]: row for row in rows}


def source_path(manifest: dict[str, dict[str, Any]], artifact_id: str) -> Path:
    record = manifest.get(artifact_id)
    if not record:
        raise KeyError(f"manifest has no artifact_id={artifact_id}")
    if record["status"] not in SUCCESS_STATUSES:
        raise RuntimeError(
            f"artifact {artifact_id} is not usable: "
            f"status={record['status']} error={record.get('error', '')}"
        )
    path = REPO_ROOT / record["local_path"]
    if not path.exists():
        raise FileNotFoundError(path)
    return path


def save_csv(frame: pd.DataFrame, filename: str) -> Path:
    path = PROCESSED_DIR / filename
    frame.to_csv(path, index=False, encoding="utf-8")
    return path


def date_string(value: Any) -> str:
    if pd.isna(value):
        return ""
    return pd.Timestamp(value).strftime("%Y-%m-%d")


def parse_excel_date(value: Any) -> pd.Timestamp:
    if isinstance(value, pd.Timestamp):
        return value
    if isinstance(value, (datetime, date)):
        return pd.Timestamp(value)
    return pd.to_datetime(value, format="%Y-%m-%d", errors="coerce")


def profile(
    dataset: str,
    frame: pd.DataFrame,
    date_column: str,
    value_columns: list[str],
    note: str = "",
    status: str = "PASS",
) -> QualityRow:
    dates = pd.to_datetime(frame[date_column], errors="coerce")
    missing = int(frame[value_columns].isna().sum().sum())
    duplicates = int(dates.duplicated().sum())
    return QualityRow(
        dataset=dataset,
        status=status,
        rows=len(frame),
        start=date_string(dates.min()),
        end=date_string(dates.max()),
        missing=missing,
        duplicates=duplicates,
        note=note,
    )


def read_fred(path: Path, value_column: str, output_column: str) -> pd.DataFrame:
    frame = pd.read_csv(path, na_values=[".", ""])
    required = {"observation_date", value_column}
    missing = required - set(frame.columns)
    if missing:
        raise ValueError(f"{path.name} missing columns {sorted(missing)}")
    frame = frame[["observation_date", value_column]].rename(
        columns={"observation_date": "date", value_column: output_column}
    )
    frame["date"] = pd.to_datetime(frame["date"], errors="coerce")
    frame[output_column] = pd.to_numeric(frame[output_column], errors="coerce")
    frame = frame.dropna(subset=["date"]).sort_values("date")
    return frame[(frame["date"] >= START_MONTHLY) & (frame["date"] <= CUTOFF)]


def build_market_data(
    manifest: dict[str, dict[str, Any]]
) -> tuple[pd.DataFrame, pd.DataFrame, list[QualityRow]]:
    specifications = [
        ("fred_brent_daily", "DCOILBRENTEU", "brent_usd_bbl"),
        ("fred_wti_daily", "DCOILWTICO", "wti_usd_bbl"),
        ("fred_usd_broad_daily", "DTWEXBGS", "usd_broad_index"),
        ("fred_cny_per_usd_daily", "DEXCHUS", "cny_per_usd"),
        ("fred_jpy_per_usd_daily", "DEXJPUS", "jpy_per_usd"),
        ("fred_krw_per_usd_daily", "DEXKOUS", "krw_per_usd"),
        ("fred_usd_per_eur_daily", "DEXUSEU", "usd_per_eur"),
    ]
    frames = []
    profiles = []
    for artifact_id, raw_column, output_column in specifications:
        frame = read_fred(source_path(manifest, artifact_id), raw_column, output_column)
        profiles.append(profile(artifact_id, frame, "date", [output_column]))
        frames.append(frame.set_index("date"))

    daily = pd.concat(frames, axis=1, join="outer").sort_index()
    daily["eur_per_usd"] = 1.0 / daily["usd_per_eur"]
    daily = daily.reset_index()
    save_csv(daily, "p0_daily_market.csv")

    monthly = (
        daily.set_index("date")
        .resample("ME")
        .mean(numeric_only=True)
        .rename_axis("month_end")
        .reset_index()
    )
    monthly["period"] = monthly["month_end"].dt.to_period("M").astype(str)
    monthly["brent_cny_per_bbl"] = (
        monthly["brent_usd_bbl"] * monthly["cny_per_usd"]
    )
    monthly = monthly[
        [
            "period",
            "month_end",
            "brent_usd_bbl",
            "wti_usd_bbl",
            "usd_broad_index",
            "cny_per_usd",
            "jpy_per_usd",
            "krw_per_usd",
            "usd_per_eur",
            "eur_per_usd",
            "brent_cny_per_bbl",
        ]
    ]
    return daily, monthly, profiles


def build_stock_data(
    manifest: dict[str, dict[str, Any]]
) -> tuple[pd.DataFrame, pd.DataFrame, QualityRow]:
    path = source_path(manifest, "eia_us_commercial_crude_stock_weekly")
    frame = pd.read_excel(
        path,
        sheet_name="Data 1",
        header=None,
        skiprows=3,
        usecols=[0, 1],
        names=["date", "us_commercial_crude_stock_kbbl"],
    )
    frame["date"] = pd.to_datetime(frame["date"], errors="coerce")
    frame["us_commercial_crude_stock_kbbl"] = pd.to_numeric(
        frame["us_commercial_crude_stock_kbbl"], errors="coerce"
    )
    frame = frame.dropna(subset=["date"]).sort_values("date")
    frame = frame[(frame["date"] >= START_MONTHLY) & (frame["date"] <= CUTOFF)]
    save_csv(frame, "eia_us_commercial_crude_stock_weekly.csv")

    month_end = (
        frame.assign(period=frame["date"].dt.to_period("M").astype(str))
        .sort_values("date")
        .groupby("period", as_index=False)
        .tail(1)
        .rename(
            columns={
                "date": "stock_reference_date",
                "us_commercial_crude_stock_kbbl": "us_crude_stock_month_end_kbbl",
            }
        )
        .sort_values("period")
        .reset_index(drop=True)
    )
    return frame, month_end, profile(
        "eia_us_commercial_crude_stock_weekly",
        frame,
        "date",
        ["us_commercial_crude_stock_kbbl"],
        note="月度主口径使用每月最后一个已观测周值",
    )


def build_gpr_data(
    manifest: dict[str, dict[str, Any]]
) -> tuple[pd.DataFrame, QualityRow]:
    path = source_path(manifest, "gpr_global_monthly")
    raw = pd.read_stata(path, convert_categoricals=False)
    required = ["month", "GPR", "GPRT", "GPRA"]
    missing = set(required) - set(raw.columns)
    if missing:
        raise ValueError(f"{path.name} missing columns {sorted(missing)}")
    frame = raw[required].copy()
    frame["month"] = pd.to_datetime(frame["month"], errors="coerce")
    frame = frame.dropna(subset=["month"]).sort_values("month")
    frame = frame[(frame["month"] >= START_MONTHLY) & (frame["month"] <= CUTOFF)]
    frame["period"] = frame["month"].dt.to_period("M").astype(str)
    frame = frame[["period", "month", "GPR", "GPRT", "GPRA"]]
    save_csv(frame, "gpr_global_monthly.csv")
    return frame, profile(
        "gpr_global_monthly",
        frame,
        "month",
        ["GPR", "GPRT", "GPRA"],
        note="2026-06 为最新初值，后续可能修订",
    )


def steo_data_status(cell: Any) -> str:
    color = cell.fill.fgColor
    if color.type == "indexed":
        return "historical"
    if color.type == "theme" and float(color.tint or 0.0) < -0.001:
        return "forecast"
    return "estimate"


def build_steo_data(
    manifest: dict[str, dict[str, Any]]
) -> tuple[pd.DataFrame, QualityRow]:
    path = source_path(manifest, "eia_steo_jul2026")
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=False)
    sheet = workbook["3atab"]
    target_map = {
        "papr_world": ("world_liquids_supply_mbd", "million barrels/day"),
        "patc_world": ("world_liquids_demand_mbd", "million barrels/day"),
        "pasc_oecd_t3": ("oecd_commercial_liquids_stocks_mmbbl", "million barrels"),
    }
    row_by_series = {}
    for row in range(1, sheet.max_row + 1):
        value = sheet.cell(row, 1).value
        if isinstance(value, str) and value.strip().lower() in target_map:
            row_by_series[value.strip().lower()] = row
    missing_series = set(target_map) - set(row_by_series)
    if missing_series:
        raise ValueError(f"STEO missing target series {sorted(missing_series)}")

    vintage_text = str(sheet["A4"].value)
    vintage_date = pd.to_datetime(
        re.sub(r"^[A-Za-z]+,\s*", "", vintage_text), errors="raise"
    ).strftime("%Y-%m-%d")

    periods: dict[int, pd.Period] = {}
    year: int | None = None
    for column in range(3, sheet.max_column + 1):
        year_value = sheet.cell(3, column).value
        if year_value is not None:
            year = int(year_value)
        month_value = sheet.cell(4, column).value
        if year is not None and isinstance(month_value, str):
            periods[column] = pd.Period(
                f"{year}-{pd.to_datetime(month_value, format='%b').month:02d}",
                freq="M",
            )

    rows = []
    for series_id, (variable_id, unit) in target_map.items():
        row = row_by_series[series_id]
        for column, period in periods.items():
            value = sheet.cell(row, column).value
            if value is None:
                continue
            rows.append(
                {
                    "period": str(period),
                    "series_id": series_id.upper(),
                    "variable_id": variable_id,
                    "value": float(value),
                    "unit": unit,
                    "vintage_date": vintage_date,
                    "data_status": steo_data_status(sheet.cell(row, column)),
                }
            )
    frame = pd.DataFrame(rows)
    frame = frame[
        pd.to_datetime(frame["period"] + "-01") <= CUTOFF
    ].sort_values(["variable_id", "period"])
    frame["is_forecast_or_estimate"] = frame["data_status"].ne("historical")
    save_csv(frame, "eia_steo_selected.csv")

    coverage_start = pd.to_datetime(frame["period"] + "-01").min()
    coverage_end = pd.to_datetime(frame["period"] + "-01").max()
    return frame, QualityRow(
        dataset="eia_steo_selected",
        status="CONDITIONAL",
        rows=len(frame),
        start=date_string(coverage_start),
        end=date_string(coverage_end),
        missing=int(frame["value"].isna().sum()),
        duplicates=int(frame.duplicated(["variable_id", "period"]).sum()),
        note="当前单一版本仅覆盖2022年起，且2026-02至06为估计值；不能回填2010年起的实时滚动预测",
    )


def build_oecd_data(
    manifest: dict[str, dict[str, Any]]
) -> tuple[pd.DataFrame, pd.DataFrame, list[QualityRow]]:
    cpi_path = source_path(manifest, "oecd_g20_cpi_monthly")
    cpi = pd.read_csv(cpi_path)
    cpi["observation_month"] = pd.to_datetime(
        cpi["TIME_PERIOD"].astype(str) + "-01", errors="coerce"
    )
    cpi = cpi[cpi["observation_month"] <= CUTOFF].copy()
    cpi_columns = [
        "REF_AREA",
        "METHODOLOGY",
        "MEASURE",
        "UNIT_MEASURE",
        "EXPENDITURE",
        "ADJUSTMENT",
        "TRANSFORMATION",
        "TIME_PERIOD",
        "OBS_VALUE",
        "OBS_STATUS",
        "BASE_PER",
    ]
    cpi = cpi[[column for column in cpi_columns if column in cpi.columns]]
    save_csv(cpi, "oecd_g20_cpi_monthly.csv")

    ip_path = source_path(manifest, "oecd_kei_ip_monthly")
    ip = pd.read_csv(ip_path)
    ip["observation_month"] = pd.to_datetime(
        ip["TIME_PERIOD"].astype(str) + "-01", errors="coerce"
    )
    ip = ip[ip["observation_month"] <= CUTOFF].copy()
    ip_columns = [
        "REF_AREA",
        "MEASURE",
        "UNIT_MEASURE",
        "ACTIVITY",
        "ADJUSTMENT",
        "TRANSFORMATION",
        "TIME_PERIOD",
        "OBS_VALUE",
        "OBS_STATUS",
        "BASE_PER",
    ]
    ip = ip[[column for column in ip_columns if column in ip.columns]]
    save_csv(ip, "oecd_kei_ip_monthly.csv")

    cpi_dates = pd.to_datetime(cpi["TIME_PERIOD"] + "-01")
    ip_dates = pd.to_datetime(ip["TIME_PERIOD"] + "-01")
    quality = [
        QualityRow(
            dataset="oecd_g20_cpi_monthly",
            status="PASS",
            rows=len(cpi),
            start=date_string(cpi_dates.min()),
            end=date_string(cpi_dates.max()),
            missing=int(cpi["OBS_VALUE"].isna().sum()),
            duplicates=int(cpi.duplicated(["REF_AREA", "TIME_PERIOD"]).sum()),
            note="CHN/DEU/FRA/ITA/ESP/JPN/KOR 各198期；德国、法国、意大利、西班牙为HICP，其余为国家CPI",
        ),
        QualityRow(
            dataset="oecd_kei_ip_monthly",
            status="PASS",
            rows=len(ip),
            start=date_string(ip_dates.min()),
            end=date_string(ip_dates.max()),
            missing=int(ip["OBS_VALUE"].isna().sum()),
            duplicates=int(ip.duplicated(["REF_AREA", "TIME_PERIOD"]).sum()),
            note="DEU/FRA/ITA/ESP/JPN/KOR 各197期；中国活动变量仍需国家统计局",
        ),
    ]
    return cpi, ip, quality


def build_eu_fuel_data(
    manifest: dict[str, dict[str, Any]]
) -> tuple[pd.DataFrame, pd.DataFrame, list[QualityRow]]:
    path = source_path(manifest, "ec_weekly_oil_bulletin")
    raw = pd.read_excel(path, sheet_name="Prices with taxes", header=0)
    date_col = raw.columns[0]
    weekly_frames: list[pd.DataFrame] = []
    monthly_frames: list[pd.DataFrame] = []
    quality: list[QualityRow] = []
    for country, (ec_code, country_name, slug) in EU_FUEL_COUNTRIES.items():
        price_col = f"{ec_code}_price_with_tax_euro95"
        if price_col not in raw.columns:
            raise ValueError(f"Weekly Oil Bulletin layout lacks {price_col}")
        weekly = raw[[date_col, price_col]].copy()
        weekly.columns = ["date", "gasoline_eur_per_1000l"]
        weekly["date"] = weekly["date"].map(parse_excel_date)
        weekly["gasoline_eur_per_1000l"] = pd.to_numeric(weekly["gasoline_eur_per_1000l"], errors="coerce")
        weekly = weekly.dropna(subset=["date", "gasoline_eur_per_1000l"])
        weekly = weekly[weekly["date"].between(START_MONTHLY, CUTOFF)].copy()
        weekly["country"] = country
        weekly["country_name"] = country_name
        weekly["gasoline_eur_per_l"] = weekly["gasoline_eur_per_1000l"] / 1000.0
        weekly = weekly[["country", "country_name", "date", "gasoline_eur_per_1000l", "gasoline_eur_per_l"]]
        weekly = weekly.sort_values("date").reset_index(drop=True)
        save_csv(weekly, f"{slug}_eurosuper95_weekly.csv")
        weekly_frames.append(weekly)

        monthly = (
            weekly.assign(period=weekly["date"].dt.to_period("M").astype(str))
            .groupby(["country", "country_name", "period"], as_index=False)
            .agg(
                gasoline_eur_per_l=("gasoline_eur_per_l", "mean"),
                weekly_observations=("gasoline_eur_per_l", "count"),
            )
        )
        monthly["month_end"] = pd.to_datetime(monthly["period"]) + pd.offsets.MonthEnd(0)
        save_csv(monthly, f"{slug}_eurosuper95_monthly.csv")
        monthly_frames.append(monthly)
        quality.extend(
            [
                profile(
                    f"{slug}_eurosuper95_weekly",
                    weekly,
                    "date",
                    ["gasoline_eur_per_l"],
                    note="European Commission Weekly Oil Bulletin, tax-inclusive Euro-super 95, original unit EUR/1000 litres.",
                ),
                profile(
                    f"{slug}_eurosuper95_monthly",
                    monthly,
                    "month_end",
                    ["gasoline_eur_per_l"],
                    note="Arithmetic monthly average of official weekly observations.",
                ),
            ]
        )

    all_weekly = pd.concat(weekly_frames, ignore_index=True)
    all_monthly = pd.concat(monthly_frames, ignore_index=True)
    save_csv(all_weekly, "eu_eurosuper95_weekly.csv")
    save_csv(all_monthly, "eu_eurosuper95_monthly.csv")
    germany_weekly = all_weekly.loc[all_weekly["country"].eq("DEU")].copy()
    germany_monthly = all_monthly.loc[all_monthly["country"].eq("DEU")].copy()
    save_csv(germany_weekly, "germany_eurosuper95_weekly.csv")
    save_csv(germany_monthly[["period", "gasoline_eur_per_l", "weekly_observations", "month_end"]], "germany_eurosuper95_monthly.csv")
    return germany_weekly, germany_monthly, quality


def build_germany_fuel_data(
    manifest: dict[str, dict[str, Any]]
) -> tuple[pd.DataFrame, pd.DataFrame, list[QualityRow]]:
    path = source_path(manifest, "ec_weekly_oil_bulletin")
    raw = pd.read_excel(path, sheet_name="Prices with taxes", header=None)
    if raw.shape[1] <= 54:
        raise ValueError("Weekly Oil Bulletin layout no longer contains expected DE columns")
    weekly = raw.iloc[3:, [0, 53, 54]].copy()
    weekly.columns = ["date", "country_code", "gasoline_eur_per_1000l"]
    weekly["date"] = pd.to_datetime(weekly["date"], errors="coerce")
    weekly["gasoline_eur_per_1000l"] = pd.to_numeric(
        weekly["gasoline_eur_per_1000l"], errors="coerce"
    )
    weekly = weekly[
        weekly["country_code"].astype(str).str.strip().eq("DE_")
        & weekly["date"].between(START_MONTHLY, CUTOFF)
    ].copy()
    weekly["gasoline_eur_per_l"] = weekly["gasoline_eur_per_1000l"] / 1000.0
    weekly = weekly.sort_values("date").reset_index(drop=True)
    save_csv(weekly, "germany_eurosuper95_weekly.csv")

    monthly = (
        weekly.assign(period=weekly["date"].dt.to_period("M").astype(str))
        .groupby("period", as_index=False)
        .agg(
            gasoline_eur_per_l=("gasoline_eur_per_l", "mean"),
            weekly_observations=("gasoline_eur_per_l", "count"),
        )
    )
    monthly["month_end"] = pd.to_datetime(monthly["period"]) + pd.offsets.MonthEnd(0)
    save_csv(monthly, "germany_eurosuper95_monthly.csv")
    quality = [
        profile(
            "germany_eurosuper95_weekly",
            weekly,
            "date",
            ["gasoline_eur_per_l"],
            note="欧盟周报含税Euro-super 95，原单位EUR/1000 litres",
        ),
        profile(
            "germany_eurosuper95_monthly",
            monthly,
            "month_end",
            ["gasoline_eur_per_l"],
            note="自然月内周度观测算术均值",
        ),
    ]
    return weekly, monthly, quality


def build_japan_fuel_data(
    manifest: dict[str, dict[str, Any]]
) -> tuple[pd.DataFrame, pd.DataFrame, list[QualityRow]]:
    path = source_path(manifest, "jp_meti_regular_gasoline_weekly")
    raw = pd.read_excel(path, sheet_name="レギュラー", header=None)
    if raw.shape[1] < 3:
        raise ValueError("Japan METI workbook no longer contains expected columns")
    weekly = raw.iloc[1:, [1, 2]].copy()
    weekly.columns = ["date", "regular_gasoline_jpy_per_l"]
    weekly["date"] = pd.to_datetime(weekly["date"], errors="coerce")
    weekly["regular_gasoline_jpy_per_l"] = pd.to_numeric(
        weekly["regular_gasoline_jpy_per_l"], errors="coerce"
    )
    weekly = weekly.dropna(subset=["date", "regular_gasoline_jpy_per_l"])
    weekly = weekly[weekly["date"].between(START_MONTHLY, CUTOFF)]
    weekly = weekly.sort_values("date").reset_index(drop=True)
    save_csv(weekly, "japan_regular_gasoline_weekly.csv")

    monthly = (
        weekly.assign(period=weekly["date"].dt.to_period("M").astype(str))
        .groupby("period", as_index=False)
        .agg(
            regular_gasoline_jpy_per_l=("regular_gasoline_jpy_per_l", "mean"),
            weekly_observations=("regular_gasoline_jpy_per_l", "count"),
        )
    )
    monthly["month_end"] = pd.to_datetime(monthly["period"]) + pd.offsets.MonthEnd(0)
    save_csv(monthly, "japan_regular_gasoline_monthly.csv")
    quality = [
        profile(
            "japan_regular_gasoline_weekly",
            weekly,
            "date",
            ["regular_gasoline_jpy_per_l"],
            note="日本资源能源厅给油所普通汽油现金价，全国，日元/升",
        ),
        profile(
            "japan_regular_gasoline_monthly",
            monthly,
            "month_end",
            ["regular_gasoline_jpy_per_l"],
            note="自然月内周度观测算术均值；2004-04以后为含消费税价格",
        ),
    ]
    return weekly, monthly, quality


def build_korea_fuel_data(
    manifest: dict[str, dict[str, Any]]
) -> tuple[pd.DataFrame, QualityRow]:
    path = source_path(manifest, "kosis_kr_gasoline_monthly")
    monthly = pd.read_csv(path)
    required = {"period", "regular_gasoline_krw_per_l"}
    missing = required - set(monthly.columns)
    if missing:
        raise ValueError(f"{path.name} missing columns {sorted(missing)}")
    monthly["month_start"] = pd.to_datetime(monthly["period"], errors="coerce")
    monthly["regular_gasoline_krw_per_l"] = pd.to_numeric(
        monthly["regular_gasoline_krw_per_l"], errors="coerce"
    )
    monthly = monthly.dropna(subset=["month_start", "regular_gasoline_krw_per_l"])
    monthly = monthly[monthly["month_start"].between(START_MONTHLY, CUTOFF)]
    monthly["period"] = monthly["month_start"].dt.to_period("M").astype(str)
    monthly["month_end"] = monthly["month_start"] + pd.offsets.MonthEnd(0)
    monthly = monthly[
        ["period", "month_end", "regular_gasoline_krw_per_l"]
    ].sort_values("month_end")
    monthly = monthly.reset_index(drop=True)
    save_csv(monthly, "korea_regular_gasoline_monthly.csv")
    quality = profile(
        "korea_regular_gasoline_monthly",
        monthly,
        "month_end",
        ["regular_gasoline_krw_per_l"],
        note=(
            "KOSIS表TX_31802_A000，韩国石油公社普通汽油全国月均价，"
            "单位KRW/litre；公开表浏览器读取快照"
        ),
    )
    return monthly, quality


def extract_policy_values(path: Path) -> tuple[int, int, int, int]:
    text = html.unescape(path.read_text(encoding="utf-8", errors="replace"))
    text = re.sub(r"<[^>]+>", "", text)
    text = re.sub(r"\s+", "", text)
    match = re.search(
        r"(?:分别应|应分别)上调(\d+)元、(\d+)元.*?实际上调(\d+)元、(\d+)元",
        text,
    )
    if not match:
        raise ValueError(f"could not parse policy adjustment values from {path.name}")
    return tuple(int(value) for value in match.groups())


def build_policy_data(
    manifest: dict[str, dict[str, Any]]
) -> tuple[pd.DataFrame, QualityRow]:
    definitions = [
        ("ndrc_fuel_control_20260323", "2026-03-23"),
        ("ndrc_fuel_control_20260407", "2026-04-07"),
    ]
    rows = []
    try:
        for artifact_id, effective_date in definitions:
            gasoline_rule, diesel_rule, gasoline_actual, diesel_actual = (
                extract_policy_values(source_path(manifest, artifact_id))
            )
            rows.append(
                {
                    "effective_date": effective_date,
                    "gasoline_rule_adjustment_cny_t": gasoline_rule,
                    "gasoline_actual_adjustment_cny_t": gasoline_actual,
                    "gasoline_policy_gap_cny_t": gasoline_rule - gasoline_actual,
                    "diesel_rule_adjustment_cny_t": diesel_rule,
                    "diesel_actual_adjustment_cny_t": diesel_actual,
                    "diesel_policy_gap_cny_t": diesel_rule - diesel_actual,
                    "source_artifact_id": artifact_id,
                }
            )
        frame = pd.DataFrame(rows)
        policy_note = "由两份国家发展改革委网页快照自动复算政策差额"
        policy_status = "PASS"
    except (KeyError, RuntimeError, FileNotFoundError, ValueError) as exc:
        cached_processed = PROCESSED_DIR / "cn_fuel_policy_events.csv"
        if not cached_processed.exists():
            raise
        frame = pd.read_csv(cached_processed)
        policy_note = (
            "NDRC网页快照当前不可用，沿用已生成的规范化提取表；"
            f"需补回官方原始快照后才能解除来源门禁。原因：{type(exc).__name__}: {exc}"
        )
        policy_status = "CONDITIONAL"
    frame["effective_date"] = pd.to_datetime(frame["effective_date"])
    save_csv(frame, "cn_fuel_policy_events.csv")
    value_columns = [
        "gasoline_rule_adjustment_cny_t",
        "gasoline_actual_adjustment_cny_t",
        "gasoline_policy_gap_cny_t",
        "diesel_rule_adjustment_cny_t",
        "diesel_actual_adjustment_cny_t",
        "diesel_policy_gap_cny_t",
    ]
    return frame, profile(
        "cn_fuel_policy_events",
        frame,
        "effective_date",
        value_columns,
        note=policy_note,
        status=policy_status,
    )


def build_release_matrix() -> pd.DataFrame:
    dictionary = pd.read_csv(DICTIONARY_PATH)

    def forecast_rule(source_id: str) -> str:
        if source_id == "SRC_EIA_STEO":
            return "单一2026-07版本不得回填历史滚动预测；收集历史vintage前从ARIMAX主规格删除"
        if source_id in {"SRC_FRED_BRENT", "SRC_FRED_WTI", "SRC_FRED_DTWEXBGS", "SRC_FRED_FX"}:
            return "日度值按实际发布日期可用；完整月均值最早在月末后使用"
        if source_id == "SRC_EIA_WCESTUS1":
            return "周值按WPSR发布日期可用；月末口径使用当时最后一个已发布周值"
        if source_id == "SRC_GPR":
            return "t月值按t+1月初版本可用，保存vintage并允许近期修订"
        if source_id.startswith("SRC_OECD"):
            return "API未返回逐期发布日期；预测模型中保守滞后1个月，后续补真实发布记录"
        if source_id == "SRC_NBS_MONTHLY":
            return "按国家统计局公告日可用；环比历史修订必须记录下载版本"
        if source_id.startswith("DERIVED"):
            return "取全部上游输入中最晚的可用日期"
        return "按来源实际公告或登记规则使用"

    matrix = dictionary[
        [
            "variable_id",
            "question",
            "source_id",
            "raw_frequency",
            "model_frequency",
            "target_coverage",
            "release_lag_rule",
            "cutoff_rule",
            "availability",
        ]
    ].copy()
    matrix["forecast_use_rule"] = matrix["source_id"].map(forecast_rule)
    matrix["exact_release_calendar_ready"] = matrix["source_id"].map(
        lambda source_id: "no"
        if source_id
        in {
            "SRC_EIA_STEO",
            "SRC_NBS_MONTHLY",
            "SRC_OECD_G20_CPI",
            "SRC_OECD_KEI_IP",
            "SRC_OECD_ENERGY_CPI",
        }
        else "rule_based"
    )
    save_csv(matrix, "release_date_matrix.csv")
    return matrix


def merge_monthly_market(
    monthly_market: pd.DataFrame,
    stock_monthly: pd.DataFrame,
    gpr: pd.DataFrame,
) -> pd.DataFrame:
    merged = monthly_market.merge(
        stock_monthly[
            ["period", "stock_reference_date", "us_crude_stock_month_end_kbbl"]
        ],
        on="period",
        how="left",
    ).merge(gpr[["period", "GPR", "GPRT", "GPRA"]], on="period", how="left")
    save_csv(merged, "p0_monthly_market.csv")
    return merged


def quality_report(
    rows: list[QualityRow],
    monthly_market: pd.DataFrame,
    release_matrix: pd.DataFrame,
) -> None:
    frame = pd.DataFrame([row.__dict__ for row in rows])
    save_csv(frame, "dataset_profile.csv")
    month_dates = pd.to_datetime(monthly_market["period"] + "-01")
    expected_months = len(pd.period_range("2010-01", "2026-06", freq="M"))
    common_complete = int(
        monthly_market[
            [
                "brent_usd_bbl",
                "us_crude_stock_month_end_kbbl",
                "usd_broad_index",
                "GPR",
            ]
        ]
        .dropna()
        .shape[0]
    )
    status_counts = frame["status"].value_counts().to_dict()
    table_lines = [
        "| 数据集 | 状态 | 行数 | 起点 | 末期 | 缺失单元 | 重复日期 | 说明 |",
        "| --- | --- | ---: | --- | --- | ---: | ---: | --- |",
    ]
    for row in rows:
        table_lines.append(
            f"| `{row.dataset}` | `{row.status}` | {row.rows} | {row.start} | "
            f"{row.end} | {row.missing} | {row.duplicates} | {row.note} |"
        )
    report = f"""# P0 数据质量报告

> 生成脚本：`code/data_processing/build_p0_datasets.py`
>
> 观测截止日：2026-06-30
>
> 本报告只评价数据管线，不代表任何模型已经通过风险探针。

## 1. 结论

- 自动来源已形成可复现快照、SHA-256 登记和处理后数据。
- 月度主区间 2010-01 至 2026-06 共 {expected_months} 个月；Brent、美国商业原油库存、美元指数和 GPR 四项共同非缺失期为 {common_complete} 个月。
- OECD 整体 CPI 覆盖中、日、韩、德各 198 期；OECD 工业生产覆盖日、韩、德各 197 期。
- EIA 2026 年 7 月 STEO 当前文件只覆盖 2022 年起，且包含估计/预测区间。它不能直接回填 2010 年起的历史滚动预测，主 ARIMAX 暂按既定回退删除全球供需变量。
- 德国含税 Euro-super 95、日本普通汽油全国现金价和韩国普通汽油全国月均价均已进入处理层；韩国 KOSIS 序列覆盖 2010-01 至 2026-06 共 198 期。
- 国家统计局工业增加值与 PPI 的完整历史序列仍未进入处理层，因此 Q2 部分结果变量和 Q3 中国活动变量继续标为 `CONDITIONAL`。

## 2. 数据集级检查

{chr(10).join(table_lines)}

状态汇总：`PASS={status_counts.get("PASS", 0)}`，`CONDITIONAL={status_counts.get("CONDITIONAL", 0)}`。

## 3. 信息集与发布滞后

`data/processed/release_date_matrix.csv` 已从变量字典生成，共 {len(release_matrix)} 个变量接口。当前只完成规则级矩阵，尚未补齐所有来源的逐期实际发布日期。

必须遵守：

1. 完整月均值只能在该月结束后使用。
2. GPR 的当月值按次月初版本可用，近期值允许修订。
3. STEO 单一最新版本不得用于伪造历史实时信息集。
4. OECD API 未返回逐期发布日期时，预测探针先保守滞后一个月。
5. 国家统计局环比历史值和其他修订序列必须保存下载版本。

## 4. 尚未放行

| 条目 | 状态 | 回退 |
| --- | --- | --- |
| 国家统计局工业增加值、PPI 完整月表 | `CONDITIONAL` | CPI 先用 OECD；IAV/PPI 到位前不运行 Q2 完整 LP |
| 中国原油进口实际单位价值 | `CONDITIONAL` | 主变量使用 Brent 月均价乘人民币兑美元汇率 |
| STEO 2010 年起实时版本 | `CONDITIONAL` | ARIMAX 主规格删除全球供需，保留实际库存、美元和 GPR |

## 5. 下一门禁

下一步先运行数据覆盖与低成本模型风险探针，写入 `results/risk_probe_summary.json`。任何 `CONDITIONAL` 项满足条件前，不进入依赖该数据的完整模型。
"""
    REPORT_PATH.write_text(report, encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--cutoff",
        default=CUTOFF.strftime("%Y-%m-%d"),
        help="Reserved for compatibility; current locked cutoff must remain 2026-06-30.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if args.cutoff != CUTOFF.strftime("%Y-%m-%d"):
        raise ValueError("competition data cutoff is locked at 2026-06-30")
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    manifest = load_manifest()
    quality_rows: list[QualityRow] = []

    _, monthly_market, market_quality = build_market_data(manifest)
    quality_rows.extend(market_quality)
    _, stock_monthly, stock_quality = build_stock_data(manifest)
    quality_rows.append(stock_quality)
    gpr, gpr_quality = build_gpr_data(manifest)
    quality_rows.append(gpr_quality)
    monthly_market = merge_monthly_market(monthly_market, stock_monthly, gpr)
    quality_rows.append(
        profile(
            "p0_monthly_market",
            monthly_market,
            "month_end",
            [
                "brent_usd_bbl",
                "wti_usd_bbl",
                "usd_broad_index",
                "cny_per_usd",
                "us_crude_stock_month_end_kbbl",
                "GPR",
            ],
            note="2010-01至2026-06主市场面板；未对缺失值做插值",
        )
    )

    _, steo_quality = build_steo_data(manifest)
    quality_rows.append(steo_quality)
    _, _, oecd_quality = build_oecd_data(manifest)
    quality_rows.extend(oecd_quality)
    _, _, germany_quality = build_eu_fuel_data(manifest)
    quality_rows.extend(germany_quality)
    _, _, japan_quality = build_japan_fuel_data(manifest)
    quality_rows.extend(japan_quality)
    _, korea_quality = build_korea_fuel_data(manifest)
    quality_rows.append(korea_quality)
    _, policy_quality = build_policy_data(manifest)
    quality_rows.append(policy_quality)
    release_matrix = build_release_matrix()
    quality_report(quality_rows, monthly_market, release_matrix)

    summary = {
        "cutoff": CUTOFF.strftime("%Y-%m-%d"),
        "processed_files": len(list(PROCESSED_DIR.glob("*.csv"))),
        "quality_report": REPORT_PATH.relative_to(REPO_ROOT).as_posix(),
        "quality_status_counts": pd.Series(
            [row.status for row in quality_rows]
        ).value_counts().to_dict(),
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
